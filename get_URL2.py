import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries


def extract_hyperlinks(file_path, cell_range):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        min_col, min_row, max_col, max_row = range_boundaries(cell_range)

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.hyperlink:
                    ws.cell(row=cell.row, column=cell.column + 1, value=cell.hyperlink.target)

        wb.save(file_path)
        print(f'超連結已提取並存入 {file_path} 中的右側欄位')

    except FileNotFoundError:
        print(f"錯誤: 找不到檔案 {file_path}")
    except OSError as e:
        print(f"OSError: {e}")
        print(f"檔案路徑可能無效或包含特殊字元: {file_path}")
    except Exception as e:
        print(f"發生未預期的錯誤: {e}")


# 用戶輸入 Excel 檔案名稱和範圍
file_path = input("請輸入 Excel 檔案名稱: ").strip()

# 移除檔案路徑前後的空白字元，這有助於移除隱藏的控制字元
file_path = file_path.strip()

print(f"您輸入的檔案路徑 (清理後): {file_path}") # 打印清理後的檔案路徑，方便debug

cell_range = input("請輸入包含超連結的範圍 (如 A1:A10): ")

extract_hyperlinks(file_path, cell_range)