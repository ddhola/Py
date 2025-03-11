import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries


def extract_hyperlinks(file_path, cell_range):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.hyperlink:
                ws.cell(row=cell.row, column=cell.column + 1, value=cell.hyperlink.target)

    wb.save(file_path)
    print(f'超連結已提取並存入 {file_path} 中的右側欄位')


# 用戶輸入 Excel 檔案名稱和範圍
file_path = input("請輸入 Excel 檔案名稱: ")
cell_range = input("請輸入包含超連結的範圍 (如 A1:A10): ")

extract_hyperlinks(file_path, cell_range)
